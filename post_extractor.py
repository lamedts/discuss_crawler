# -*- coding: utf8 -*-

#################################################################
# post_extractor was written in its entirety by Dan Stuart      #
# And is distributed under Creative Commons by-nc-sa            #
# (Attribution, non-commercial, share-alike).                   #
# Contact me at dstu@umich.edu or drestuart@gmail.com           #
# With any questions.  5/17/2010                                #
###                                                             #
### @   A post extractor for discuss.hk base on post_extractor  #
###         which created by Dan Stuart                         #
### @   Download openpyxl from write and read execl for python  #
###     command-> python -m pip install openpyxl                #
### @   Run this program                                        #
###     Command-> python post_extractor.py                      #
#################################################################

import urllib
import sys
from parse_post import *

import codecs
out = codecs.getwriter('utf-8')(sys.stdout)

# denpendence 
from openpyxl import Workbook
from openpyxl.styles import Alignment 

def big5_utf8(strObj):
    try:
        returnStr = strObj.decode('big5').encode('utf8')
    except:
        returnStr = 'encoding problem'

    return returnStr


def main():
    #if len(sys.argv) == 1:
    #    print "Usage: python post_extractor.py http://yoursite.com/forum firstpage# lastpage# outfile.txt"
    #    exit(0)

    sys.argv.pop(0)
    try:
        threadid = sys.argv.pop(0)
    except IndexError:
        threadid = raw_input("Enter the url of the threadid(tid=?, e.g. 24942698): ") or "24942698"
        #e.g. "http://www.discuss.com.hk/viewthread.php?tid=24574577"
        url_start = "http://www.discuss.com.hk/viewthread.php?tid=" + threadid

    try:
        first_page = int(sys.argv.pop(0))
    except IndexError:
        first_page = int(raw_input("Enter the page number you would like to start with: ") or "1")

    try:
        last_page = int(sys.argv.pop(0))
    except IndexError:
        last_page = int(raw_input("Enter the page number of the last page: ") or "1") + 1

    try:
        outfilename = sys.argv.pop(0)
        outfilename = outfilename + ".xlsx"
    except IndexError:
        outfilename = raw_input("Enter the name of the output execl(w/o filetype): ") or "1"
        outfilename = outfilename + ".xlsx"

    print "%s  %d  %d  %s" % (url_start, first_page, last_page, outfilename)
    print "\nterminal charset: " + sys.stdin.encoding + "\n"

    urls = []

    for i in range(first_page, last_page):
        urls.append( url_start + '&page=' + str(i) )

    idx = 1
    wb = Workbook()
    ws = wb.active
    arg = Alignment(wrap_text=True)

    id_header = ws.cell(row = idx, column = 1)
    id_header.value = "Post ID"

    user_header = ws.cell(row = idx, column = 2)
    user_header.value = "User"
    ws.column_dimensions['B'].width = 20

    timestamp_header = ws.cell(row = idx, column = 3)
    timestamp_header.value = "Timestamp"
    ws.column_dimensions['C'].width = 30

    date_header = ws.cell(row = idx, column = 4)
    date_header.value = "Date"

    hour_header = ws.cell(row = idx, column = 5)
    hour_header.value = "Hour"

    quote_header = ws.cell(row = idx, column = 6)
    quote_header.value = "Quote"
    quote_header.alignment = arg
    ws.column_dimensions['F'].width = 100

    reply_header = ws.cell(row = idx, column = 7)
    reply_header.value = "Reply"
    reply_header.alignment = arg
    ws.column_dimensions['G'].width = 100

    for url in urls:
        req = urllib.urlopen(url)
        html_string = req.read()
        users, posts, times, postids, title, dates, hours, quotes = parse_post(html_string)

        if len(users) != len(posts):
            print "Length mismatch!  url: " + url
            print "len(users) = " + str(len(users)) + "  len(posts): " + str(len(posts))
            exit(1)
        
        for i in range(len(users)):
            user = big5_utf8(users[i])
            time = big5_utf8(times[i])
            date = big5_utf8(dates[i])
            hour = hours[i]
            id = postids[i]
            post = ''
            quote = ''
            for j in range(len(posts[i])):
                post += big5_utf8(str(posts[i][j]))
            quote += big5_utf8(str(quotes[i]))
            idx = idx + 1
            id_cell = ws.cell(row = idx, column = 1)
            id_cell.value = id
            user_cell = ws.cell(row = idx, column = 2)
            user_cell.value = user
            time_cell = ws.cell(row = idx, column = 3)
            time_cell.value = time
            date_cell = ws.cell(row = idx, column = 4)
            date_cell.value = date
            hour_cell = ws.cell(row = idx, column = 5)
            hour_cell.value = hour
            quote_cell = ws.cell(row = idx, column = 6)
            quote_cell.value = quote
            quote_cell.alignment = arg
            post_cell = ws.cell(row = idx, column = 7)
            post_cell.value = post
            post_cell.alignment = arg
        print "Finished"

    wb.save(outfilename)
    print "Created"


'''
start the program
'''
main()

